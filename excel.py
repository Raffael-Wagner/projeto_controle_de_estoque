from PySide2.QtWidgets import QFileDialog, QMessageBox
from PySide2.QtWidgets import QTreeWidgetItem
import openpyxl
import pandas as pd

#***Função para selecionar um arquivo
def open_file_dialog(main_window):
    options = QFileDialog.Options()
    fileName, _ = QFileDialog.getOpenFileName(main_window, "Abrir Arquivo Excel", "", "Excel Files (*.xlsx);", options=options)
    if fileName:
        main_window.txt_file.setText(fileName)  

#***Função para iniciar a importação do arquivo Excel***
def import_excel(main_window):
    file_path = main_window.txt_file.text()  
    if file_path:
        main_window.excel_file_path = file_path
        read_excel(main_window, file_path)
        QMessageBox.information(main_window, "Importação Concluída", "Importação realizada com sucesso.")

def read_excel(main_window, file_path):
    pasta_de_trabalho = openpyxl.load_workbook(file_path)
    folha = pasta_de_trabalho.active

    linhas = folha.max_row
    colunas = folha.max_column

    for linha in range(2, linhas + 1):
        item = QTreeWidgetItem()

        for coluna in range(1, colunas + 1):
            valor_celula = folha.cell(row=linha, column=coluna).value
            item.setText(coluna - 1, str(valor_celula))
        main_window.tw_estoque.addTopLevelItem(item)
#***Função gerar excel do estoque criado de maneira manual***
def gerar_excel_estoque_manual(main_window):
    headers = []
    for column in range(main_window.tw_estoque.columnCount()):
        headers.append(main_window.tw_estoque.headerItem().text(column))
        
    data = []
    for row in range(main_window.tw_estoque.topLevelItemCount()):
        item_data = []
        for column in range(main_window.tw_estoque.columnCount()):
            item_data.append(main_window.tw_estoque.topLevelItem(row).text(column))
        data.append(item_data)

    df = pd.DataFrame(data, columns=headers)

    options = QFileDialog.Options()
    file_path, _ = QFileDialog.getSaveFileName(main_window, "Salvar arquivo Excel", "", "Excel Files (*.xlsx)", options=options)
    if file_path:
        df.to_excel(file_path, index=False)
        QMessageBox.information(main_window, "Operação concluída", f"O arquivo foi salvo com sucesso!")

def update_excel(main_window):
    if not main_window.excel_file_path:
        return

    workbook = openpyxl.load_workbook(main_window.excel_file_path)
    sheet = workbook.active

    sheet.delete_rows(2, sheet.max_row)

    for indice_linha_atual in range(main_window.tw_estoque.topLevelItemCount()):
        item = main_window.tw_estoque.topLevelItem(indice_linha_atual)
        for indice_coluna_atual in range(item.columnCount()):
            sheet.cell(row=indice_linha_atual + 2, column=indice_coluna_atual + 1).value = item.text(indice_coluna_atual)

    workbook.save(main_window.excel_file_path)